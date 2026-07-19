"""Viewer reader plugin for Excel spreadsheets (the ``documents`` extra).

Provides a preview of ``.xlsx``/``.xlsm`` workbooks in the built-in
viewer's table view. Uses openpyxl's ``read_only`` streaming mode so the
whole workbook is never loaded into memory.

Registers no extensions when the ``openpyxl`` package (part of the
``documents`` extra) isn't installed, per the guard convention documented in
``plugins/__init__.py``.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from linux_commander.vfs import FileSystem, VfsPath

if TYPE_CHECKING:
    from linux_commander.plugins import ViewDocument

try:
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.styles import Alignment, Font, PatternFill

    VIEW_EXTENSIONS: tuple[str, ...] = (".xlsx", ".xlsm")
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    Font = PatternFill = Alignment = None  # type: ignore[assignment,misc]
    VIEW_EXTENSIONS = ()


def _cell_text(value: object) -> str:
    return "" if value is None else str(value)


def _cell_format(cell) -> dict:
    """Extract basic formatting info from a cell."""
    fmt = {}
    if cell.font:
        if cell.font.bold:
            fmt["bold"] = True
        if cell.font.italic:
            fmt["italic"] = True
        if cell.font.color and cell.font.color.rgb:
            fmt["color"] = cell.font.color.rgb
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        fmt["bg"] = cell.fill.start_color.rgb
    if cell.alignment:
        if cell.alignment.horizontal:
            fmt["align"] = cell.alignment.horizontal
    return fmt


def read_document(host_fs: FileSystem, path: VfsPath) -> ViewDocument:
    """Read the workbook as a table preview with sheet metadata.

    Materializes the workbook to a real file (openpyxl needs to seek within
    the zip container) and streams rows with ``read_only=True`` so a large
    workbook is not fully loaded. Returns sheet names for the selector and
    formatting hints for the first sheet.
    """
    from linux_commander.plugins import MAX_PREVIEW_ROWS, ViewDocument, cleanup_temp, materialize

    real_path = materialize(host_fs, path)
    try:
        wb = openpyxl.load_workbook(real_path, read_only=True, data_only=True)
        try:
            sheet_names = wb.sheetnames
            # For preview, use the first sheet but include all sheet names
            sheet = wb.worksheets[0]
            rows: list[list[str]] = []
            formats: list[list[dict]] = []
            truncated = False
            for i, row in enumerate(sheet.iter_rows(values_only=False)):
                if i >= MAX_PREVIEW_ROWS:
                    truncated = True
                    break
                row_values = [_cell_text(c.value) for c in row]
                row_formats = [_cell_format(c) for c in row]
                rows.append(row_values)
                formats.append(row_formats)

            # Include sheet names for selector
            meta = {
                "sheets": sheet_names,
                "active_sheet": 0,
                "formats": formats,
            }
            return ViewDocument(kind="table", rows=rows, truncated=truncated, meta=meta)
        finally:
            wb.close()
    finally:
        cleanup_temp(real_path)
