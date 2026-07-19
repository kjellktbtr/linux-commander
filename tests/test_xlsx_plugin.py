"""Tests for the xlsx viewer reader plugin (the ``documents`` extra).

Fixtures are built with openpyxl directly, so these run anywhere the optional
dependency is installed. Skipped entirely when ``openpyxl`` isn't installed.
"""

from __future__ import annotations

from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

import openpyxl  # noqa: E402

from linux_commander.plugins import viewer_plugin_for_name  # noqa: E402
from linux_commander.plugins.xlsx_plugin import read_document  # noqa: E402
from linux_commander.vfs import LocalFileSystem, VfsPath  # noqa: E402

_FS = LocalFileSystem()


def _local_vpath(path: Path) -> VfsPath:
    return _FS.from_path(path)


def _make_xlsx(path: Path, rows: list[list[object]]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# viewer_plugin_for_name discovery
# ---------------------------------------------------------------------------


def test_xlsx_extension_is_discovered() -> None:
    mod = viewer_plugin_for_name("book.xlsx")
    assert mod is not None
    assert mod.read_document is read_document


def test_xlsm_extension_is_discovered() -> None:
    assert viewer_plugin_for_name("macro.xlsm") is not None


# ---------------------------------------------------------------------------
# read_document
# ---------------------------------------------------------------------------


def test_read_document_returns_header_and_rows(tmp_path: Path) -> None:
    xlsx = _make_xlsx(tmp_path / "book.xlsx", [["name", "role"], ["Alice", "Dev"]])
    doc = read_document(_FS, _local_vpath(xlsx))
    assert doc.kind == "table"
    assert doc.rows[0] == ["name", "role"]
    assert doc.rows[1] == ["Alice", "Dev"]
    assert doc.truncated is False


def test_read_document_only_uses_first_sheet(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["a", "b"])
    ws2 = wb.create_sheet("Second")
    ws2.append(["x", "y"])
    xlsx = tmp_path / "multi.xlsx"
    wb.save(str(xlsx))

    doc = read_document(_FS, _local_vpath(xlsx))
    assert doc.rows == [["a", "b"]]


def test_read_document_converts_none_cells_to_empty_string(tmp_path: Path) -> None:
    xlsx = _make_xlsx(tmp_path / "sparse.xlsx", [["a", "b"], ["only-a", None]])
    doc = read_document(_FS, _local_vpath(xlsx))
    assert doc.rows[1] == ["only-a", ""]


def test_read_document_caps_rows_and_marks_truncated(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr("linux_commander.plugins.MAX_PREVIEW_ROWS", 5)
    rows = [["col"]] + [[str(i)] for i in range(15)]
    xlsx = _make_xlsx(tmp_path / "big.xlsx", rows)
    doc = read_document(_FS, _local_vpath(xlsx))
    assert doc.truncated is True
    assert len(doc.rows) == 5


def test_read_document_not_truncated_when_under_cap(tmp_path: Path) -> None:
    rows = [["col"]] + [[str(i)] for i in range(10)]
    xlsx = _make_xlsx(tmp_path / "small.xlsx", rows)
    doc = read_document(_FS, _local_vpath(xlsx))
    assert doc.truncated is False
    assert len(doc.rows) == 11
